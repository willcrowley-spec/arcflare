"""Document parsing with format-specific handlers.

Routes documents to the appropriate parser based on MIME type.
Supports PDF, DOCX, PPTX, XLSX, and plain text.
"""
import logging
import mimetypes
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class ParsedElement:
    """A single element extracted from a document."""
    text: str
    element_type: str  # "title", "narrative", "list_item", "table", "key_value", "code_block"
    page_number: int | None = None
    metadata: dict | None = None


def parse_file(file_path: str | Path) -> list[ParsedElement]:
    """Parse a document file and return structured elements."""
    path = Path(file_path)
    mime_type, _ = mimetypes.guess_type(str(path))

    if mime_type is None:
        mime_type = _guess_from_extension(path.suffix.lower())

    logger.info("parsing file=%s mime=%s", path.name, mime_type)

    if mime_type in ("application/pdf",):
        return _parse_with_unstructured(path)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ):
        return _parse_with_unstructured(path)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "application/vnd.ms-powerpoint",
    ):
        return _parse_with_unstructured(path)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return _parse_excel(path)
    elif mime_type in ("application/json",):
        return _parse_json(path)
    elif mime_type and mime_type.startswith("text/"):
        return _parse_text(path)
    else:
        logger.warning("unsupported_mime_type file=%s mime=%s, falling back to text", path.name, mime_type)
        return _parse_text(path)


def _guess_from_extension(ext: str) -> str:
    mapping = {
        ".pdf": "application/pdf",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".doc": "application/msword",
        ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".json": "application/json",
        ".txt": "text/plain",
        ".csv": "text/csv",
        ".md": "text/markdown",
    }
    return mapping.get(ext, "application/octet-stream")


def _parse_with_unstructured(path: Path) -> list[ParsedElement]:
    """Parse PDF, DOCX, PPTX using unstructured library."""
    try:
        from unstructured.partition.auto import partition
        raw_elements = partition(filename=str(path))
    except ImportError:
        logger.warning("unstructured not installed, falling back to text extraction")
        return _parse_text(path)
    except Exception as e:
        logger.error("unstructured_parse_failed file=%s error=%s", path.name, e)
        return _parse_text(path)

    elements: list[ParsedElement] = []
    for el in raw_elements:
        text = str(el).strip()
        if not text:
            continue

        el_type = type(el).__name__.lower()
        if "title" in el_type or "header" in el_type:
            element_type = "title"
        elif "list" in el_type:
            element_type = "list_item"
        elif "table" in el_type:
            element_type = "table"
        else:
            element_type = "narrative"

        page = getattr(el.metadata, "page_number", None) if hasattr(el, "metadata") else None

        elements.append(ParsedElement(
            text=text,
            element_type=element_type,
            page_number=page,
        ))

    logger.info("parsed_with_unstructured file=%s elements=%d", path.name, len(elements))
    return elements


def _parse_excel(path: Path) -> list[ParsedElement]:
    """Parse Excel files using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed, falling back to text")
        return [ParsedElement(text=f"[Excel file: {path.name}]", element_type="narrative")]

    elements: list[ParsedElement] = []
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheetnames = list(wb.sheetnames)

    for sheet_name in sheetnames:
        ws = wb[sheet_name]
        elements.append(ParsedElement(
            text=f"Sheet: {sheet_name}",
            element_type="title",
        ))

        rows_text = []
        for row in ws.iter_rows(values_only=True):
            cell_values = [str(c) if c is not None else "" for c in row]
            if any(v.strip() for v in cell_values):
                rows_text.append(" | ".join(cell_values))

        if rows_text:
            elements.append(ParsedElement(
                text="\n".join(rows_text),
                element_type="table",
            ))

    wb.close()
    logger.info("parsed_excel file=%s sheets=%d elements=%d", path.name, len(sheetnames), len(elements))
    return elements


def _parse_json(path: Path) -> list[ParsedElement]:
    """Parse JSON files into structured elements."""
    import json
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logger.warning("json_parse_failed file=%s error=%s", path.name, e)
        return _parse_text(path)

    text = json.dumps(data, indent=2, default=str)
    if len(text) > 50_000:
        text = text[:50_000] + "\n... [truncated]"

    return [ParsedElement(
        text=text,
        element_type="key_value",
        metadata={"format": "json"},
    )]


def _parse_text(path: Path) -> list[ParsedElement]:
    """Parse plain text / markdown / CSV files."""
    try:
        text = path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = path.read_text(encoding="latin-1")
    except Exception as e:
        logger.error("text_read_failed file=%s error=%s", path.name, e)
        return []

    if not text.strip():
        return []

    return [ParsedElement(
        text=text,
        element_type="narrative",
    )]
